import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Feuil1']
    PRICE_COLUMN = 3
    CORRECTED_PRICE_COLUMN = 6
    sheet.cell(1, CORRECTED_PRICE_COLUMN).value = 'corrected price'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, PRICE_COLUMN)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, CORRECTED_PRICE_COLUMN)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'g2')

    workbook.save(filename)
